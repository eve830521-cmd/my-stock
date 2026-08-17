import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "역DCF_계산기"

# Styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
label_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

ws['A1'] = "■ JW생명과학 역DCF (Reverse DCF) 계산기"
ws['A1'].font = Font(bold=True, size=14)

ws['A3'] = "[입력 데이터]"
ws['A3'].font = Font(bold=True)

inputs = [
    ("현재 주가 (원)", 11364, '#,##0'),
    ("총 발행주식수 (주)", 15836092, '#,##0'),
    ("기준 FCF (억원) ★", 400, '#,##0'),
    ("할인율 (WACC)", 0.10, '0.0%'),
    ("영구성장률 (Terminal)", 0.02, '0.0%')
]

for i, (label, val, fmt) in enumerate(inputs):
    row = i + 4
    ws[f'A{row}'] = label
    ws[f'A{row}'].fill = label_fill
    ws[f'B{row}'] = val
    ws[f'B{row}'].number_format = fmt

ws['A10'] = "[계산 결과]"
ws['A10'].font = Font(bold=True)

ws['A11'] = "현재 시가총액 (억원)"
ws['A11'].fill = label_fill
ws['B11'] = "=B4*B5/100000000"
ws['B11'].number_format = '#,##0'
ws['B11'].font = Font(bold=True)

ws['A13'] = "★ 시장 내재 성장률 (Implied Growth Rate)"
ws['A13'].font = Font(bold=True, color="C00000")
ws['A13'].fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
ws['B13'] = "=INDEX(F:F, MATCH(MIN(N:N), N:N, 0))"
ws['B13'].number_format = '0.00%'
ws['B13'].font = Font(bold=True, color="C00000", size=12)
ws['B13'].fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

ws['A14'] = "👉 현재 시가총액을 정당화하기 위해 향후 5년간 매년 필요한 FCF 성장률입니다."
ws['A14'].font = Font(italic=True, color="595959")

# Data Table for iteration (Columns F to N)
headers = ["g (성장률)", "FCF 1", "FCF 2", "FCF 3", "FCF 4", "FCF 5", "Terminal Value", "PV (계산된 시총)", "차이 (Abs Diff)"]
for col_idx, text in enumerate(headers, start=6):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = text
    cell.font = Font(color="808080")

# Fill from -30% to +50% in steps of 0.1% (801 steps)
start_g = -0.30
for i in range(801):
    row = i + 2
    g_val = start_g + (i * 0.001)
    ws[f'F{row}'] = g_val
    ws[f'G{row}'] = f"=\\*(1+\{row})"
    ws[f'H{row}'] = f"=\{row}*(1+\{row})"
    ws[f'I{row}'] = f"=\{row}*(1+\{row})"
    ws[f'J{row}'] = f"=\{row}*(1+\{row})"
    ws[f'K{row}'] = f"=\{row}*(1+\{row})"
    ws[f'L{row}'] = f"=\{row}*(1+\\)/(\\-\\)"
    ws[f'M{row}'] = f"=\{row}/(1+\\) + \{row}/((1+\\)^2) + \{row}/((1+\\)^3) + \{row}/((1+\\)^4) + \{row}/((1+\\)^5) + \{row}/((1+\\)^5)"
    ws[f'N{row}'] = f"=ABS(\{row}-\\)"

ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 20

# Hide columns F to N so it looks clean to the user
for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
    ws.column_dimensions[col].hidden = True

wb.save("E:/antigravity-work/my-skill/Stock_Research/JW생명과학/3_Final_Reports/JW생명과학_Reverse_DCF.xlsx")
print("Excel file created successfully.")
